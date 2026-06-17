"""
LocalHost monthly Excel template generator.

A minimal, modern, monochrome workbook a household admin fills in once a
month and uploads to the LocalHost app.

Tabs:
- Dashboard : header + roster (member name + email = identifier) + KPIs
- Meals     : pivot grids (days x members) for each meal slot, plus daily totals
- Costs     : market shopping, fixed costs, other expenses, electricity reading

Design:
- Black / white / grey palette only.
- Hidden gridlines, refined typography, subtle horizontal dividers.
- Live SUM formulas feed the dashboard KPIs.
- Hidden Lists sheet powers all dropdowns.
- A hidden Schema sheet documents how each cell maps to the database.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ────────────────────────────────────────────────────────────────────
# Configuration

OUTPUT_PATH = Path("public/assets/LocalHost_Monthly_Template_Simple.xlsx")
TEMPLATE_VERSION = "1.0"

MEMBER_SLOTS = 8                          # how many member rows in the roster
MEAL_SLOTS = ("Lunch", "Dinner")          # default meal slots in the system
MEAL_DAYS = 31
MARKET_ROWS = 24
FIXED_ROWS = 12
EXPENSE_ROWS = 24

# Monochrome palette
INK = "121212"
CHARCOAL = "2B2B2B"
GRAPHITE = "555555"
STEEL = "8C8C8C"
SILVER = "BFBFBF"
MIST = "E8E8E8"
FOG = "F2F2F2"
SNOW = "FAFAFA"
WHITE = "FFFFFF"


# ────────────────────────────────────────────────────────────────────
# Style helpers

def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def font(*, color: str = INK, bold: bool = False, italic: bool = False, size: int = 10, name: str = "Calibri") -> Font:
    return Font(name=name, color=color, bold=bold, italic=italic, size=size)


def edge(color: str = SILVER, style: str = "thin") -> Side:
    return Side(border_style=style, color=color)


def left(indent: int = 1, wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=indent)


def center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def right(indent: int = 1) -> Alignment:
    return Alignment(horizontal="right", vertical="center", indent=indent)


def style_range(ws, cell_range, *, fill_=None, font_=None, alignment=None, border=None):
    target = ws[cell_range]
    # openpyxl returns a single Cell for a 1-cell range, a tuple of Cells for a
    # row/column range, and a tuple of tuples for a 2-D range. Normalize.
    from openpyxl.cell.cell import Cell

    if isinstance(target, Cell):
        rows = ((target,),)
    elif target and isinstance(target[0], Cell):
        rows = (target,)
    else:
        rows = target

    for row in rows:
        for cell in row:
            if fill_ is not None:
                cell.fill = fill_
            if font_ is not None:
                cell.font = font_
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def hairline(top: bool = False, bottom: bool = False, color: str = MIST) -> Border:
    return Border(
        top=edge(color) if top else None,
        bottom=edge(color) if bottom else None,
    )


def brand_header(ws, last_col: str, title: str, subtitle: str) -> None:
    """Black bar at the top of every visible sheet."""

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 10

    ws.merge_cells(f"B2:{last_col}2")
    ws["B2"] = title
    ws["B2"].font = font(color=WHITE, bold=True, size=20)
    ws["B2"].fill = fill(INK)
    ws["B2"].alignment = left(indent=1)

    ws.merge_cells(f"B3:{last_col}3")
    ws["B3"] = subtitle
    ws["B3"].font = font(color="D9D9D9", italic=True, size=10)
    ws["B3"].fill = fill(INK)
    ws["B3"].alignment = left(indent=1)


def section_header(ws, range_: str, label: str) -> None:
    """Section banner inside a sheet."""

    first = range_.split(":")[0]
    row = int("".join(c for c in first if c.isdigit()))
    ws.row_dimensions[row].height = 22
    ws.merge_cells(range_)
    ws[first] = label
    ws[first].font = font(color=WHITE, bold=True, size=11)
    ws[first].fill = fill(CHARCOAL)
    ws[first].alignment = left(indent=1)


def table_header(ws, row: int, headers: list[str], start_col: int = 2) -> None:
    ws.row_dimensions[row].height = 26
    for i, name in enumerate(headers):
        col = get_column_letter(start_col + i)
        c = ws[f"{col}{row}"]
        c.value = name
        c.font = font(color=WHITE, bold=True)
        c.fill = fill(INK)
        c.alignment = center(wrap=True)


def zebra_rows(ws, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    for r in range(first_row, last_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(WHITE if r % 2 == 0 else SNOW)
            cell.border = hairline(bottom=True)


# ────────────────────────────────────────────────────────────────────
# Dashboard

def build_dashboard(ws) -> dict[str, str]:
    set_widths(
        ws,
        {
            "A": 4,
            "B": 6,
            "C": 26,
            "D": 30,
            "E": 16,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 4,
        },
    )

    brand_header(ws, "K", "LOCALHOST", f"Monthly data sheet  ·  v{TEMPLATE_VERSION}  ·  fill, save, upload")

    # ── Meta inputs (underlined-field style) ──────────────────────
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 22

    def meta(row_label: int, row_value: int, label_cell: str, value_range: str, label: str) -> None:
        ws[label_cell] = label
        ws[label_cell].font = font(color=GRAPHITE, bold=True, size=9)
        ws[label_cell].alignment = Alignment(horizontal="left", vertical="bottom")
        ws.merge_cells(value_range)
        first = value_range.split(":")[0]
        ws[first].font = font(size=12, bold=True)
        ws[first].alignment = left()
        style_range(ws, value_range, border=Border(bottom=edge(STEEL)))

    meta(5, 6, "B5", "B6:D6", "APARTMENT NAME")
    meta(5, 6, "E5", "E6:G6", "MONTH")
    meta(5, 6, "H5", "H6:K6", "PREPARED BY")

    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 22
    meta(8, 9, "B8", "B9:D9", "REGISTRATION ID")
    meta(8, 9, "E8", "E9:G9", "SUBMISSION DATE")
    meta(8, 9, "H8", "H9:K9", "CONTACT (PHONE / EMAIL)")

    # ── Member roster ────────────────────────────────────────────
    roster_section_row = 11
    section_header(ws, f"B{roster_section_row}:K{roster_section_row}", "MEMBER ROSTER")

    caption_row = roster_section_row + 1
    ws.row_dimensions[caption_row].height = 16
    ws.merge_cells(f"B{caption_row}:K{caption_row}")
    ws[f"B{caption_row}"] = (
        "EMAIL is how the system identifies each member. Use the exact email already registered in the app. "
        "The order here defines Member 1, Member 2 … used in the Meals tab."
    )
    ws[f"B{caption_row}"].font = font(color=GRAPHITE, italic=True, size=9)
    ws[f"B{caption_row}"].alignment = left(wrap=True)
    ws[f"B{caption_row}"].fill = fill(SNOW)

    header_row = caption_row + 1
    table_header(ws, header_row, ["#", "FULL NAME", "EMAIL  (identifier)", "STATUS", "ROLE", "PHONE", "NOTES"], start_col=2)
    # merge NOTES across two columns for breathing room
    ws.merge_cells(f"H{header_row}:K{header_row}")
    ws[f"H{header_row}"].value = "NOTES"
    ws[f"H{header_row}"].font = font(color=WHITE, bold=True)
    ws[f"H{header_row}"].fill = fill(INK)
    ws[f"H{header_row}"].alignment = center()

    roster_first = header_row + 1
    roster_last = roster_first + MEMBER_SLOTS - 1
    for i in range(MEMBER_SLOTS):
        r = roster_first + i
        ws.row_dimensions[r].height = 22
        # # column
        ws[f"B{r}"] = f"{i + 1:02d}"
        ws[f"B{r}"].font = font(color=STEEL, bold=True)
        ws[f"B{r}"].alignment = center()
        ws[f"B{r}"].fill = fill(FOG)
        # editable cells
        for col in ("C", "D", "E", "F", "G"):
            cell = ws[f"{col}{r}"]
            cell.alignment = left() if col in ("C", "D", "G") else center()
            cell.fill = fill(WHITE if r % 2 == 0 else SNOW)
            cell.border = hairline(bottom=True)
        ws.merge_cells(f"H{r}:K{r}")
        ws[f"H{r}"].alignment = left()
        ws[f"H{r}"].fill = fill(WHITE if r % 2 == 0 else SNOW)
        for c in ("H", "I", "J", "K"):
            ws[f"{c}{r}"].border = hairline(bottom=True)

    # ── KPI cards ────────────────────────────────────────────────
    kpi_row_title = roster_last + 2
    kpi_row_value = kpi_row_title + 1
    kpi_row_caption = kpi_row_value + 1
    ws.row_dimensions[kpi_row_title].height = 18
    ws.row_dimensions[kpi_row_value].height = 38
    ws.row_dimensions[kpi_row_caption].height = 18

    cards = [
        ("FIXED COSTS",   "=IFERROR('Costs'!D27,0)", "#,##0.00"),
        ("OTHER EXPENSES","=IFERROR('Costs'!E55,0)", "#,##0.00"),
        ("MARKET TOTAL",  "=IFERROR('Costs'!E12,0)", "#,##0.00"),
        ("TOTAL MEALS",   "=IFERROR('Meals'!N79,0)", "#,##0.0"),
    ]
    starts = [("B", "D"), ("E", "F"), ("G", "H"), ("I", "K")]
    for (label, formula, fmt), (col, col2) in zip(cards, starts):
        title_range = f"{col}{kpi_row_title}:{col2}{kpi_row_title}"
        value_range = f"{col}{kpi_row_value}:{col2}{kpi_row_value}"
        caption_range = f"{col}{kpi_row_caption}:{col2}{kpi_row_caption}"

        ws.merge_cells(title_range)
        ws[f"{col}{kpi_row_title}"] = label
        ws[f"{col}{kpi_row_title}"].font = font(color=GRAPHITE, bold=True, size=9)
        ws[f"{col}{kpi_row_title}"].alignment = left()
        ws[f"{col}{kpi_row_title}"].fill = fill(FOG)

        ws.merge_cells(value_range)
        ws[f"{col}{kpi_row_value}"] = formula
        ws[f"{col}{kpi_row_value}"].font = font(color=INK, bold=True, size=22)
        ws[f"{col}{kpi_row_value}"].alignment = left()
        ws[f"{col}{kpi_row_value}"].fill = fill(FOG)
        ws[f"{col}{kpi_row_value}"].number_format = fmt

        ws.merge_cells(caption_range)
        ws[f"{col}{kpi_row_caption}"] = "auto-calculated  ·  live"
        ws[f"{col}{kpi_row_caption}"].font = font(color=STEEL, italic=True, size=8)
        ws[f"{col}{kpi_row_caption}"].alignment = left()
        ws[f"{col}{kpi_row_caption}"].fill = fill(FOG)

        for rng in (title_range, value_range, caption_range):
            style_range(ws, rng, border=Border(top=edge(MIST), bottom=edge(MIST)))

    # Footer help
    foot_row = kpi_row_caption + 2
    ws.row_dimensions[foot_row].height = 18
    ws.merge_cells(f"B{foot_row}:K{foot_row}")
    ws[f"B{foot_row}"] = "Need help?  Fill the Roster first, then Meals, then Costs.  Cells in white are editable."
    ws[f"B{foot_row}"].font = font(color=STEEL, italic=True, size=9)
    ws[f"B{foot_row}"].alignment = left()

    ws.freeze_panes = f"A{roster_section_row}"

    return {
        "roster_first": str(roster_first),
        "roster_last": str(roster_last),
    }


# ────────────────────────────────────────────────────────────────────
# Meals

def build_meals(ws, roster_first: int, roster_last: int) -> str:
    set_widths(
        ws,
        {
            "A": 4,
            "B": 8,
            **{get_column_letter(3 + i): 13 for i in range(MEMBER_SLOTS)},
            get_column_letter(3 + MEMBER_SLOTS): 14,
            get_column_letter(4 + MEMBER_SLOTS): 4,
        },
    )

    last_col_letter = get_column_letter(3 + MEMBER_SLOTS)
    brand_header(ws, last_col_letter, "MEALS", f"One sheet per meal slot  ·  {len(MEAL_SLOTS)} slots")

    grand_total_refs: list[str] = []
    next_row = 5

    for slot_name in MEAL_SLOTS:
        # Section banner
        section_header(ws, f"B{next_row}:{last_col_letter}{next_row}", slot_name.upper())
        caption_row = next_row + 1
        ws.row_dimensions[caption_row].height = 16
        ws.merge_cells(f"B{caption_row}:{last_col_letter}{caption_row}")
        ws[f"B{caption_row}"] = "Enter meal count per member per day (e.g. 1, 1.5 with guest, 0 if skipped)."
        ws[f"B{caption_row}"].font = font(color=GRAPHITE, italic=True, size=9)
        ws[f"B{caption_row}"].alignment = left(wrap=True)
        ws[f"B{caption_row}"].fill = fill(SNOW)

        # Column headers: DAY + member labels (linked to roster name)
        header_row = caption_row + 1
        ws.row_dimensions[header_row].height = 36
        ws[f"B{header_row}"] = "DAY"
        ws[f"B{header_row}"].font = font(color=WHITE, bold=True)
        ws[f"B{header_row}"].fill = fill(INK)
        ws[f"B{header_row}"].alignment = center()

        for i in range(MEMBER_SLOTS):
            col = get_column_letter(3 + i)
            roster_row = roster_first + i
            # Show "M01 — <linked name>" so the user can see at a glance which roster row.
            ws[f"{col}{header_row}"] = (
                f"=\"M{i + 1:02d} — \" & IF('Dashboard'!C{roster_row}=\"\",\"(empty)\",'Dashboard'!C{roster_row})"
            )
            ws[f"{col}{header_row}"].font = font(color=WHITE, bold=True, size=10)
            ws[f"{col}{header_row}"].fill = fill(INK)
            ws[f"{col}{header_row}"].alignment = center(wrap=True)

        total_col = get_column_letter(3 + MEMBER_SLOTS)
        ws[f"{total_col}{header_row}"] = "DAY TOTAL"
        ws[f"{total_col}{header_row}"].font = font(color=WHITE, bold=True)
        ws[f"{total_col}{header_row}"].fill = fill(CHARCOAL)
        ws[f"{total_col}{header_row}"].alignment = center()

        first_data = header_row + 1
        last_data = first_data + MEAL_DAYS - 1
        for d in range(MEAL_DAYS):
            r = first_data + d
            ws.row_dimensions[r].height = 20
            ws[f"B{r}"] = f"{d + 1:02d}"
            ws[f"B{r}"].font = font(color=GRAPHITE, bold=True)
            ws[f"B{r}"].alignment = center()
            ws[f"B{r}"].fill = fill(FOG)
            ws[f"B{r}"].border = Border(right=edge(MIST))

            for i in range(MEMBER_SLOTS):
                col = get_column_letter(3 + i)
                cell = ws[f"{col}{r}"]
                cell.alignment = center()
                cell.number_format = "0.0;;;@"
                cell.fill = fill(WHITE if d % 2 == 0 else SNOW)
                cell.border = hairline(bottom=True)

            first_m = get_column_letter(3)
            last_m = get_column_letter(2 + MEMBER_SLOTS)
            ws[f"{total_col}{r}"] = f"=IFERROR(SUM({first_m}{r}:{last_m}{r}),0)"
            ws[f"{total_col}{r}"].font = font(color=INK, bold=True)
            ws[f"{total_col}{r}"].alignment = center()
            ws[f"{total_col}{r}"].number_format = "0.0;;;@"
            ws[f"{total_col}{r}"].fill = fill(FOG)
            ws[f"{total_col}{r}"].border = Border(left=edge(MIST), bottom=edge(MIST))

        # Slot total row
        slot_total_row = last_data + 1
        ws.row_dimensions[slot_total_row].height = 24
        ws[f"B{slot_total_row}"] = f"{slot_name.upper()} TOTAL"
        ws[f"B{slot_total_row}"].font = font(color=WHITE, bold=True)
        ws[f"B{slot_total_row}"].fill = fill(CHARCOAL)
        ws[f"B{slot_total_row}"].alignment = center()

        for i in range(MEMBER_SLOTS):
            col = get_column_letter(3 + i)
            ws[f"{col}{slot_total_row}"] = f"=IFERROR(SUM({col}{first_data}:{col}{last_data}),0)"
            ws[f"{col}{slot_total_row}"].font = font(color=WHITE, bold=True)
            ws[f"{col}{slot_total_row}"].fill = fill(CHARCOAL)
            ws[f"{col}{slot_total_row}"].alignment = center()
            ws[f"{col}{slot_total_row}"].number_format = "0.0;;;@"

        ws[f"{total_col}{slot_total_row}"] = f"=IFERROR(SUM({total_col}{first_data}:{total_col}{last_data}),0)"
        ws[f"{total_col}{slot_total_row}"].font = font(color=WHITE, bold=True, size=12)
        ws[f"{total_col}{slot_total_row}"].fill = fill(INK)
        ws[f"{total_col}{slot_total_row}"].alignment = center()
        ws[f"{total_col}{slot_total_row}"].number_format = "0.0;;;@"

        grand_total_refs.append(f"{total_col}{slot_total_row}")
        next_row = slot_total_row + 2

    # Grand total bar covering all slots
    ws.row_dimensions[next_row].height = 28
    ws.merge_cells(f"B{next_row}:{get_column_letter(2 + MEMBER_SLOTS)}{next_row}")
    ws[f"B{next_row}"] = "ALL MEALS — MONTH GRAND TOTAL"
    ws[f"B{next_row}"].font = font(color=WHITE, bold=True, size=11)
    ws[f"B{next_row}"].fill = fill(INK)
    ws[f"B{next_row}"].alignment = left()

    grand_cell = f"{get_column_letter(3 + MEMBER_SLOTS)}{next_row}"
    ws[grand_cell] = "=" + "+".join(grand_total_refs) if grand_total_refs else "=0"
    ws[grand_cell].font = font(color=WHITE, bold=True, size=14)
    ws[grand_cell].fill = fill(INK)
    ws[grand_cell].alignment = center()
    ws[grand_cell].number_format = "0.0;;;@"

    ws.freeze_panes = "C5"

    return f"'Meals'!{grand_cell}"


# ────────────────────────────────────────────────────────────────────
# Costs (Market + Fixed + Other + Bill)

def build_costs(ws, roster_first: int, roster_last: int) -> dict[str, str]:
    set_widths(
        ws,
        {
            "A": 4,
            "B": 14,
            "C": 28,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
            "H": 28,
            "I": 4,
        },
    )

    brand_header(ws, "H", "COSTS", "Market shopping, fixed costs, monthly bill, and other expenses")

    # ── Market shopping ─────────────────────────────────────────
    section_header(ws, "B5:H5", "MEAL MARKET SHOPPING  ·  who bought what for meals")
    table_header(ws, 6, ["Date", "Item", "Quantity", "Amount", "Bought By (Email)", "Receipt #", "Notes"])
    market_first = 7
    market_last = market_first + MARKET_ROWS - 1
    zebra_rows(ws, market_first, market_last, 2, 8)
    for r in range(market_first, market_last + 1):
        ws[f"B{r}"].number_format = "yyyy-mm-dd"
        ws[f"E{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].alignment = left()
        ws[f"H{r}"].alignment = left()

    market_total_row = market_last + 1
    ws.row_dimensions[market_total_row].height = 24
    ws.merge_cells(f"B{market_total_row}:D{market_total_row}")
    ws[f"B{market_total_row}"] = "MARKET TOTAL"
    ws[f"B{market_total_row}"].font = font(color=WHITE, bold=True)
    ws[f"B{market_total_row}"].fill = fill(CHARCOAL)
    ws[f"B{market_total_row}"].alignment = left()
    ws[f"E{market_total_row}"] = f"=IFERROR(SUM(E{market_first}:E{market_last}),0)"
    ws[f"E{market_total_row}"].font = font(color=WHITE, bold=True, size=12)
    ws[f"E{market_total_row}"].fill = fill(INK)
    ws[f"E{market_total_row}"].alignment = center()
    ws[f"E{market_total_row}"].number_format = "#,##0.00"
    for col in ("F", "G", "H"):
        ws[f"{col}{market_total_row}"].fill = fill(CHARCOAL)

    # ── Fixed costs ─────────────────────────────────────────────
    fixed_section_row = market_total_row + 2
    section_header(ws, f"B{fixed_section_row}:H{fixed_section_row}", "FIXED COSTS  ·  monthly recurring amounts")
    table_header(ws, fixed_section_row + 1, ["Cost Item", "Amount", "Paid By (Email)", "Payment Method", "Reference", "Due Date", "Notes"])
    fixed_first = fixed_section_row + 2
    fixed_last = fixed_first + FIXED_ROWS - 1
    zebra_rows(ws, fixed_first, fixed_last, 2, 8)
    for r in range(fixed_first, fixed_last + 1):
        ws[f"C{r}"].number_format = "#,##0.00"
        ws[f"G{r}"].number_format = "yyyy-mm-dd"
        ws[f"B{r}"].alignment = left()
        ws[f"H{r}"].alignment = left()

    fixed_total_row = fixed_last + 1
    ws.row_dimensions[fixed_total_row].height = 24
    ws.merge_cells(f"B{fixed_total_row}:B{fixed_total_row}")
    ws[f"B{fixed_total_row}"] = "FIXED TOTAL"
    ws[f"B{fixed_total_row}"].font = font(color=WHITE, bold=True)
    ws[f"B{fixed_total_row}"].fill = fill(CHARCOAL)
    ws[f"B{fixed_total_row}"].alignment = left()
    ws[f"C{fixed_total_row}"] = f"=IFERROR(SUM(C{fixed_first}:C{fixed_last}),0)"
    ws[f"C{fixed_total_row}"].font = font(color=WHITE, bold=True, size=12)
    ws[f"C{fixed_total_row}"].fill = fill(INK)
    ws[f"C{fixed_total_row}"].alignment = center()
    ws[f"C{fixed_total_row}"].number_format = "#,##0.00"
    for col in ("D", "E", "F", "G", "H"):
        ws[f"{col}{fixed_total_row}"].fill = fill(CHARCOAL)

    # ── Monthly bill ────────────────────────────────────────────
    bill_section_row = fixed_total_row + 2
    section_header(ws, f"B{bill_section_row}:H{bill_section_row}", "MONTHLY BILL  ·  meter readings and adjustments")
    bill_header_row = bill_section_row + 1
    ws.row_dimensions[bill_header_row].height = 26
    ws[f"B{bill_header_row}"] = "ELECTRICITY READING (units)"
    ws[f"B{bill_header_row}"].font = font(color=GRAPHITE, bold=True, size=9)
    ws[f"B{bill_header_row}"].alignment = left()
    ws.merge_cells(f"C{bill_header_row}:D{bill_header_row}")
    ws[f"C{bill_header_row}"].font = font(size=12, bold=True)
    ws[f"C{bill_header_row}"].alignment = center()
    ws[f"C{bill_header_row}"].fill = fill(WHITE)
    ws[f"C{bill_header_row}"].border = Border(bottom=edge(STEEL))
    ws[f"C{bill_header_row}"].number_format = "#,##0"

    ws[f"E{bill_header_row}"] = "ELECTRICITY AMOUNT"
    ws[f"E{bill_header_row}"].font = font(color=GRAPHITE, bold=True, size=9)
    ws[f"E{bill_header_row}"].alignment = left()
    ws.merge_cells(f"F{bill_header_row}:G{bill_header_row}")
    ws[f"F{bill_header_row}"].font = font(size=12, bold=True)
    ws[f"F{bill_header_row}"].alignment = center()
    ws[f"F{bill_header_row}"].fill = fill(WHITE)
    ws[f"F{bill_header_row}"].border = Border(bottom=edge(STEEL))
    ws[f"F{bill_header_row}"].number_format = "#,##0.00"

    adj_header_row = bill_header_row + 2
    table_header(ws, adj_header_row, ["Member Email", "Type (extra/deduct)", "Label", "Amount", "Created By", "Date", "Notes"])
    adj_first = adj_header_row + 1
    adj_last = adj_first + 7
    zebra_rows(ws, adj_first, adj_last, 2, 8)
    for r in range(adj_first, adj_last + 1):
        ws[f"E{r}"].number_format = "#,##0.00"
        ws[f"G{r}"].number_format = "yyyy-mm-dd"
        ws[f"B{r}"].alignment = left()
        ws[f"D{r}"].alignment = left()
        ws[f"H{r}"].alignment = left()

    # ── Other expenses ──────────────────────────────────────────
    other_section_row = adj_last + 2
    section_header(ws, f"B{other_section_row}:H{other_section_row}", "OTHER EXPENSES  ·  one-off and per-member items")
    table_header(ws, other_section_row + 1, ["Date", "Category", "Description", "Bought By (Email)", "Amount", "Payment Method", "Notes"])
    other_first = other_section_row + 2
    other_last = other_first + EXPENSE_ROWS - 1
    zebra_rows(ws, other_first, other_last, 2, 8)
    for r in range(other_first, other_last + 1):
        ws[f"B{r}"].number_format = "yyyy-mm-dd"
        ws[f"F{r}"].number_format = "#,##0.00"
        ws[f"D{r}"].alignment = left()
        ws[f"H{r}"].alignment = left()

    other_total_row = other_last + 1
    ws.row_dimensions[other_total_row].height = 24
    ws.merge_cells(f"B{other_total_row}:E{other_total_row}")
    ws[f"B{other_total_row}"] = "OTHER EXPENSES TOTAL"
    ws[f"B{other_total_row}"].font = font(color=WHITE, bold=True)
    ws[f"B{other_total_row}"].fill = fill(CHARCOAL)
    ws[f"B{other_total_row}"].alignment = left()
    ws[f"F{other_total_row}"] = f"=IFERROR(SUM(F{other_first}:F{other_last}),0)"
    ws[f"F{other_total_row}"].font = font(color=WHITE, bold=True, size=12)
    ws[f"F{other_total_row}"].fill = fill(INK)
    ws[f"F{other_total_row}"].alignment = center()
    ws[f"F{other_total_row}"].number_format = "#,##0.00"
    for col in ("G", "H"):
        ws[f"{col}{other_total_row}"].fill = fill(CHARCOAL)

    ws.freeze_panes = "A5"

    return {
        "market_total": f"E{market_total_row}",
        "fixed_total": f"C{fixed_total_row}",
        "other_total": f"F{other_total_row}",
        "fixed_first": str(fixed_first),
        "fixed_last": str(fixed_last),
        "other_first": str(other_first),
        "other_last": str(other_last),
        "adj_first": str(adj_first),
        "adj_last": str(adj_last),
        "market_first": str(market_first),
        "market_last": str(market_last),
    }


# ────────────────────────────────────────────────────────────────────
# Lists (hidden)

def add_lists(wb) -> None:
    lists = wb.create_sheet("Lists")
    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    payment_methods = ["Cash", "Bank Transfer", "bKash", "Nagad", "Card", "Other"]
    categories = [
        "Grocery", "Utility", "Transport", "Internet",
        "Maintenance", "Cleaning", "Medicine", "Repair", "Other",
    ]
    statuses = ["Active", "Guest", "Inactive"]
    roles = ["Admin", "Bill Manager", "Member"]
    adj_types = ["extra", "deduct"]

    columns = [months, payment_methods, categories, statuses, roles, adj_types]
    for col_idx, values in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        for i, v in enumerate(values, start=1):
            lists[f"{col_letter}{i}"] = v
    lists.sheet_state = "hidden"


# ────────────────────────────────────────────────────────────────────
# Schema reference (hidden)

def add_schema_reference(wb, refs: dict[str, str]) -> None:
    sheet = wb.create_sheet("Schema")
    sheet.sheet_state = "hidden"
    sheet["A1"] = "LocalHost Monthly Template — Schema mapping"
    sheet["A1"].font = font(bold=True, size=12)
    sheet["A3"] = "Section"
    sheet["B3"] = "Sheet range"
    sheet["C3"] = "Maps to (Prisma model)"
    sheet["D3"] = "Notes"
    for cell in ("A3", "B3", "C3", "D3"):
        sheet[cell].font = font(bold=True)

    rows = [
        ("Roster",         f"Dashboard!C{refs['roster_first']}:G{refs['roster_last']}", "Member",       "Match by email (apartmentId+email is unique)"),
        ("Meals",          "Meals!C…",                                             "MealRecord",   "One sheet section per meal slot"),
        ("Market",         f"Costs!B{refs['market_first']}:H{refs['market_last']}",     "MealShopping", "Per-purchase rows"),
        ("Fixed costs",    f"Costs!B{refs['fixed_first']}:H{refs['fixed_last']}",       "FixedCost",    "Monthly recurring items"),
        ("Bill / adj",     f"Costs!B{refs['adj_first']}:H{refs['adj_last']}",           "BillAdjustment / MonthlyBill", "Electricity reading + per-member adjustments"),
        ("Other expenses", f"Costs!B{refs['other_first']}:H{refs['other_last']}",       "Expense",      "Variable monthly expenses"),
    ]
    for i, row in enumerate(rows, start=4):
        for j, val in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=val)


# ────────────────────────────────────────────────────────────────────
# Data validations

def add_validations(wb, dashboard_ws, meals_ws, costs_ws, dashboard_refs: dict[str, str], costs_refs: dict[str, str]) -> None:
    # Month dropdown on dashboard
    month_dv = DataValidation(type="list", formula1="=Lists!$A$1:$A$12", allow_blank=False)
    dashboard_ws.add_data_validation(month_dv)
    month_dv.add("E6")

    # Status / role dropdowns in roster
    status_dv = DataValidation(type="list", formula1="=Lists!$D$1:$D$3", allow_blank=True)
    role_dv = DataValidation(type="list", formula1="=Lists!$E$1:$E$3", allow_blank=True)
    dashboard_ws.add_data_validation(status_dv)
    dashboard_ws.add_data_validation(role_dv)
    rf = int(dashboard_refs["roster_first"])
    rl = int(dashboard_refs["roster_last"])
    status_dv.add(f"E{rf}:E{rl}")
    role_dv.add(f"F{rf}:F{rl}")

    # Costs validations
    pay_dv = DataValidation(type="list", formula1="=Lists!$B$1:$B$6", allow_blank=True)
    cat_dv = DataValidation(type="list", formula1="=Lists!$C$1:$C$9", allow_blank=True)
    adj_dv = DataValidation(type="list", formula1="=Lists!$F$1:$F$2", allow_blank=True)
    costs_ws.add_data_validation(pay_dv)
    costs_ws.add_data_validation(cat_dv)
    costs_ws.add_data_validation(adj_dv)

    pay_dv.add(f"E{costs_refs['fixed_first']}:E{costs_refs['fixed_last']}")
    pay_dv.add(f"G{costs_refs['other_first']}:G{costs_refs['other_last']}")
    cat_dv.add(f"C{costs_refs['other_first']}:C{costs_refs['other_last']}")
    adj_dv.add(f"C{costs_refs['adj_first']}:C{costs_refs['adj_last']}")

    # Meal counts are constrained to non-negative numbers <= 5 to catch typos
    meal_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="5", allow_blank=True)
    meals_ws.add_data_validation(meal_dv)
    last_col_letter = get_column_letter(2 + MEMBER_SLOTS)
    for offset in range(len(MEAL_SLOTS)):
        # Each slot has 31 day rows starting at row (5 + offset*(MEAL_DAYS+5)) — but simpler to apply broadly:
        # apply to whole member-data area
        pass
    # Apply to a wide range that covers both slots' data; harmless to overshoot.
    meal_dv.add(f"C5:{last_col_letter}200")


# ────────────────────────────────────────────────────────────────────
# Build

def build_workbook() -> None:
    wb = Workbook()

    dashboard = wb.active
    dashboard.title = "Dashboard"
    dashboard_refs = build_dashboard(dashboard)
    roster_first = int(dashboard_refs["roster_first"])
    roster_last = int(dashboard_refs["roster_last"])

    meals = wb.create_sheet("Meals")
    meals_grand_total = build_meals(meals, roster_first, roster_last)

    costs = wb.create_sheet("Costs")
    costs_refs = build_costs(costs, roster_first, roster_last)

    # Patch dashboard KPI formulas to actual references now that Costs/Meals exist.
    kpi_value_row = roster_last + 3  # title=roster_last+2, value=roster_last+3
    dashboard.cell(row=kpi_value_row, column=2).value  = f"='Costs'!{costs_refs['fixed_total']}"
    dashboard.cell(row=kpi_value_row, column=5).value  = f"='Costs'!{costs_refs['other_total']}"
    dashboard.cell(row=kpi_value_row, column=7).value  = f"='Costs'!{costs_refs['market_total']}"
    dashboard.cell(row=kpi_value_row, column=9).value  = f"={meals_grand_total}"

    add_lists(wb)
    add_schema_reference(
        wb,
        refs={
            "roster_first": str(roster_first),
            "roster_last": str(roster_last),
            **costs_refs,
        },
    )
    add_validations(wb, dashboard, meals, costs, dashboard_refs, costs_refs)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()

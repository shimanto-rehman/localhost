"""
Generate the LocalHost Monthly Data Import Template Excel file.
Run: python3 scripts/generate_template.py
Output: public/assets/LocalHost_Monthly_Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
import datetime

# ── Color Palette (matches LocalHost brand) ─────────────────────────────────
C = {
    'brand':       '14B8A6',   # teal-500
    'brand_dark':  '0D9488',   # teal-600
    'brand_light': '99F6E4',   # teal-200
    'sky':         '38BDF8',
    'amber':       'F59E0B',
    'rose':        'FB7185',
    'violet':      'A78BFA',
    'green':       '22C55E',
    'slate_900':   '0F172A',
    'slate_800':   '1E293B',
    'slate_700':   '334155',
    'slate_600':   '475569',
    'slate_500':   '64748B',
    'slate_400':   '94A3B8',
    'slate_300':   'CBD5E1',
    'slate_200':   'E2E8F0',
    'slate_100':   'F1F5F9',
    'slate_50':    'F8FAFC',
    'white':       'FFFFFF',
    'gold':        'FCD34D',
    'food':        'F59E0B',
    'groceries':   '22C55E',
    'utilities':   '38BDF8',
    'transport':   'A78BFA',
    'household':   'FB7185',
    'entertainment':'F472B6',
    'medical':     'EF4444',
    'other':       '94A3B8',
}

EXPENSE_CATEGORIES = ['Food', 'Groceries', 'Utilities', 'Transport',
                      'Household', 'Entertainment', 'Medical', 'Other']

MONTH_NAMES = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']

# Sample member names (users replace these)
SAMPLE_MEMBERS = ['Ali Hassan', 'Rahim Uddin', 'Karim Molla', 'Sumon Ahmed']

# ── Helpers ──────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(bold=False, size=11, color='0F172A', italic=False, name='Calibri'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides='all'):
    s = Side(style='thin', color='E2E8F0')
    m = Side(style='medium', color='94A3B8')
    if sides == 'all':
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == 'bottom':
        return Border(bottom=m)
    if sides == 'top':
        return Border(top=m)
    return Border()

def thick_border():
    t = Side(style='medium', color='14B8A6')
    return Border(left=t, right=t, top=t, bottom=t)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def cell(ws, row, col, value='', bg=None, bold=False, size=11, color='0F172A',
         h='left', v='center', wrap=False, italic=False, border=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=color)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c

def merge_cell(ws, row1, col1, row2, col2, value='', bg=None, bold=False,
               size=11, color='0F172A', h='center', v='center', wrap=False,
               italic=False, border=None):
    ws.merge_cells(start_row=row1, start_column=col1,
                   end_row=row2, end_column=col2)
    c = ws.cell(row=row1, column=col1, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if border:
        c.border = border
    return c

def section_header(ws, row, col_start, col_end, title, subtitle=''):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.fill = fill(C['brand_dark'])
    c.font = font(bold=True, size=12, color=C['white'])
    c.alignment = align(h='left', v='center')
    row_height(ws, row, 28)
    return row + 1

def col_header(ws, row, col, value):
    c = cell(ws, row, col, value, bg=C['slate_800'], bold=True,
             size=10, color=C['brand_light'], h='center', border=thin_border())
    row_height(ws, row, 24)
    return c

def data_cell(ws, row, col, value='', bg=C['white'], num_format=None):
    return cell(ws, row, col, value, bg=bg, size=10, h='left',
                border=thin_border(), num_format=num_format)

def note_cell(ws, row, col_start, col_end, text, bg=C['slate_100']):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(italic=True, size=9, color=C['slate_500'])
    c.alignment = align(h='left', v='center')
    row_height(ws, row, 18)

# ── Sheet 1: Cover ────────────────────────────────────────────────────────────

def build_cover(wb):
    ws = wb.create_sheet('📋 Instructions', 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # Set column widths
    widths = [3, 18, 32, 18, 18, 18, 3]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, get_column_letter(i), w)

    # Hero banner
    for r in range(1, 9):
        row_height(ws, r, 22)
        for c_ in range(1, 8):
            ws.cell(row=r, column=c_).fill = fill(C['slate_900'])

    row_height(ws, 3, 14)
    row_height(ws, 7, 14)

    merge_cell(ws, 4, 2, 4, 6,
               '🏠  LocalHost — Monthly Data Import Template',
               bg=C['slate_900'], bold=True, size=20, color=C['brand_light'], h='left')
    merge_cell(ws, 5, 2, 5, 6,
               'Fill in the tabs below, then import from the app → Upload Data',
               bg=C['slate_900'], bold=False, size=11, color=C['slate_400'], h='left', italic=True)
    merge_cell(ws, 6, 2, 6, 6,
               f'Template version 1.0  ·  {datetime.date.today().strftime("%B %Y")}',
               bg=C['slate_900'], bold=False, size=9, color=C['slate_600'], h='left')

    # Divider
    for c_ in range(1, 8):
        ws.cell(row=8, column=c_).fill = fill(C['brand'])

    row = 10

    # Step boxes
    steps = [
        ('1', 'Members', C['brand'],
         'Go to 👥 Members tab.\nList all active members in Column B.\nDo NOT change the header row.'),
        ('2', 'Bills', C['sky'],
         'Go to 🧾 Bills tab.\nEnter electricity bill amount.\nAdd custom cost overrides if needed.'),
        ('3', 'Meals', C['amber'],
         'Go to 🍽 Meals tab.\nMark ✓ for confirmed meals per member per day.\nFill shopping entries below.'),
        ('4', 'Expenses', C['rose'],
         'Go to 💳 Expenses tab.\nLog personal expenses with category & date.\nOne row per expense item.'),
    ]

    row_height(ws, row, 22)
    merge_cell(ws, row, 2, row, 6, 'HOW TO USE THIS TEMPLATE',
               bg=C['white'], bold=True, size=13, color=C['slate_800'], h='left')
    row += 1

    for step in steps:
        num, title, color_, desc = step
        row_height(ws, row, 14)
        # number badge
        c_ = ws.cell(row=row+1, column=2, value=step[0])
        c_.fill = fill(color_)
        c_.font = font(bold=True, size=18, color=C['white'])
        c_.alignment = align(h='center', v='center')
        ws.merge_cells(start_row=row+1, start_column=2,
                        end_row=row+3, end_column=2)

        ws.merge_cells(start_row=row+1, start_column=3,
                        end_row=row+1, end_column=6)
        t = ws.cell(row=row+1, column=3, value=step[1])
        t.font = font(bold=True, size=13, color=color_)
        t.alignment = align(h='left', v='center')
        t.fill = fill(C['slate_50'])

        ws.merge_cells(start_row=row+2, start_column=3,
                        end_row=row+3, end_column=6)
        d = ws.cell(row=row+2, column=3, value=step[3])
        d.font = font(size=10, color=C['slate_600'], italic=True)
        d.alignment = align(h='left', v='center', wrap=True)
        d.fill = fill(C['slate_50'])

        for r2 in range(row, row+5):
            row_height(ws, r2, 18)
        row_height(ws, row+1, 26)
        row_height(ws, row+2, 22)
        row_height(ws, row+3, 22)
        row += 5

    # Rules
    row += 1
    row_height(ws, row, 22)
    merge_cell(ws, row, 2, row, 6, '⚠  IMPORTANT RULES',
               bg=C['amber'] + '33', bold=True, size=11, color=C['amber'], h='left')
    row += 1

    rules = [
        '• Do NOT rename or reorder the sheet tabs.',
        '• Do NOT change the grey header rows (Row 1, Row 3 etc.).',
        '• Month Key format is YYYY-MM  (e.g. 2025-06).',
        '• All amounts are in BDT (integer, no decimals).',
        '• Member names must exactly match what is in the app.',
        '• Meal confirmed = ✓ (the letter V also works). Leave blank if not confirmed.',
        '• Expense Date format: DD/MM/YYYY  (e.g. 15/06/2025).',
    ]
    for r_text in rules:
        row_height(ws, row, 18)
        merge_cell(ws, row, 2, row, 6, r_text,
                   bg=C['white'], bold=False, size=10, color=C['slate_700'], h='left')
        row += 1

    return ws

# ── Sheet 2: Members ─────────────────────────────────────────────────────────

def build_members(wb):
    ws = wb.create_sheet('👥 Members')
    ws.sheet_view.showGridLines = False

    widths = [3, 28, 22, 22, 3]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, get_column_letter(i), w)

    # Banner
    for c_ in range(1, 6):
        ws.cell(row=1, column=c_).fill = fill(C['slate_900'])
    row_height(ws, 1, 10)

    merge_cell(ws, 2, 2, 2, 4, '👥  Members List',
               bg=C['slate_900'], bold=True, size=15, color=C['brand_light'], h='left')
    row_height(ws, 2, 34)
    merge_cell(ws, 3, 2, 3, 4,
               'List all active members below. Names must match the app exactly.',
               bg=C['brand_dark'] + '22', bold=False, size=9, color=C['slate_500'], h='left', italic=True)
    row_height(ws, 3, 18)

    # Headers
    row = 4
    row_height(ws, row, 26)
    headers = ['#', 'Member Name', 'Role (optional)', 'Notes (optional)']
    for i, h_ in enumerate(headers):
        col_header(ws, row, i + 2, h_)

    # Data rows
    for idx, name in enumerate(SAMPLE_MEMBERS, 1):
        row += 1
        row_height(ws, row, 22)
        bg = C['slate_50'] if idx % 2 == 0 else C['white']
        data_cell(ws, row, 2, idx, bg=bg)
        data_cell(ws, row, 3, name, bg=bg)
        data_cell(ws, row, 4, 'Member', bg=bg)
        data_cell(ws, row, 5, '', bg=bg)

    # Extra blank rows
    for idx in range(len(SAMPLE_MEMBERS) + 1, len(SAMPLE_MEMBERS) + 8):
        row += 1
        row_height(ws, row, 22)
        bg = C['slate_50'] if idx % 2 == 0 else C['white']
        data_cell(ws, row, 2, idx, bg=bg)
        data_cell(ws, row, 3, '', bg=bg)
        data_cell(ws, row, 4, '', bg=bg)
        data_cell(ws, row, 5, '', bg=bg)

    row += 2
    note_cell(ws, row, 2, 4, 'ℹ  Members are referenced by name in other sheets.')

    return ws

# ── Sheet 3: Bills ────────────────────────────────────────────────────────────

def build_bills(wb):
    ws = wb.create_sheet('🧾 Bills')
    ws.sheet_view.showGridLines = False

    widths = [3, 30, 18, 18, 18, 18, 3]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, get_column_letter(i), w)

    # Banner
    for c_ in range(1, 8):
        ws.cell(row=1, column=c_).fill = fill(C['slate_900'])
    row_height(ws, 1, 10)

    merge_cell(ws, 2, 2, 2, 6, '🧾  Monthly Bills',
               bg=C['slate_900'], bold=True, size=15, color=C['sky'], h='left')
    row_height(ws, 2, 34)
    merge_cell(ws, 3, 2, 3, 6,
               'Enter one value per field. All amounts in BDT (integers only).',
               bg=C['sky'] + '22', bold=False, size=9, color=C['slate_500'], h='left', italic=True)
    row_height(ws, 3, 18)

    row = 5

    # ── Section A: Month & Electricity ──
    row = section_header(ws, row, 2, 6, '  SECTION A — Month & Electricity')

    for label, placeholder, hint in [
        ('Month Key', '2025-06', 'Format: YYYY-MM'),
        ('Electricity Bill (BDT)', '0', 'Total bill as shown on meter receipt'),
    ]:
        row_height(ws, row, 24)
        cell(ws, row, 2, label, bg=C['slate_100'], bold=True, size=10,
             color=C['slate_700'], border=thin_border())
        c_ = ws.cell(row=row, column=3, value=placeholder)
        c_.fill = fill(C['white'])
        c_.font = font(size=10, color=C['sky'], bold=True)
        c_.alignment = align(h='center')
        c_.border = thin_border()
        ws.merge_cells(start_row=row, start_column=3,
                        end_row=row, end_column=4)
        cell(ws, row, 5, f'← {hint}', bg=C['slate_50'], italic=True,
             size=9, color=C['slate_400'], border=thin_border())
        row += 1

    row += 1

    # ── Section B: Fixed Costs ──
    row = section_header(ws, row, 2, 6, '  SECTION B — Fixed Costs')
    note_cell(ws, row, 2, 6, 'These are apartment-level costs split across all members.')
    row += 1

    row_height(ws, row, 26)
    for i, h_ in enumerate(['Cost Name', 'Amount (BDT)', 'Override?', 'Notes'], 2):
        col_header(ws, row, i, h_)
    row += 1

    fixed_costs = [
        ('Base House Rent', 20000),
        ('Gas Bill', 1080),
        ('Water Bill', 1000),
        ('Building Service Charge', 2000),
    ]
    for idx, (name, amount) in enumerate(fixed_costs):
        row_height(ws, row, 22)
        bg = C['slate_50'] if idx % 2 == 0 else C['white']
        cell(ws, row, 2, name, bg=bg, size=10, border=thin_border())
        c_ = ws.cell(row=row, column=3, value=amount)
        c_.fill = fill(bg)
        c_.font = font(size=10, color=C['sky'], bold=True)
        c_.alignment = align(h='center')
        c_.border = thin_border()
        c_.number_format = '#,##0'
        cell(ws, row, 4, '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 5, '', bg=bg, size=10, border=thin_border())
        row += 1

    # Blank rows for custom fixed costs
    for i in range(3):
        row_height(ws, row, 22)
        bg = C['white'] if i % 2 == 0 else C['slate_50']
        cell(ws, row, 2, '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 3, 0, bg=bg, size=10, h='center', border=thin_border())
        cell(ws, row, 4, '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 5, '', bg=bg, size=10, border=thin_border())
        row += 1

    row += 1

    # ── Section C: Optional Costs ──
    row = section_header(ws, row, 2, 6, '  SECTION C — Optional / Per-member Costs')
    note_cell(ws, row, 2, 6, 'Only members who opted-in are charged. List opted-in members in col E.')
    row += 1

    row_height(ws, row, 26)
    for i, h_ in enumerate(['Cost Name', 'Amount (BDT)', 'Opted-in Members (comma separated)', 'Notes'], 2):
        col_header(ws, row, i, h_)
    row += 1

    optional_costs = [
        ('House Maid', 2500, ', '.join(SAMPLE_MEMBERS)),
        ('Wi-Fi Bill', 800, ', '.join(SAMPLE_MEMBERS[:2])),
    ]
    for idx, (name, amount, members) in enumerate(optional_costs):
        row_height(ws, row, 22)
        bg = C['slate_50'] if idx % 2 == 0 else C['white']
        cell(ws, row, 2, name, bg=bg, size=10, border=thin_border())
        c_ = ws.cell(row=row, column=3, value=amount)
        c_.fill = fill(bg)
        c_.font = font(size=10, color=C['amber'], bold=True)
        c_.alignment = align(h='center')
        c_.border = thin_border()
        c_.number_format = '#,##0'
        cell(ws, row, 4, members, bg=bg, size=9, border=thin_border(), wrap=True)
        cell(ws, row, 5, '', bg=bg, size=10, border=thin_border())
        row += 1

    for i in range(3):
        row_height(ws, row, 22)
        bg = C['white'] if i % 2 == 0 else C['slate_50']
        cell(ws, row, 2, '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 3, 0, bg=bg, size=10, h='center', border=thin_border())
        cell(ws, row, 4, '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 5, '', bg=bg, size=10, border=thin_border())
        row += 1

    row += 1

    # ── Section D: Adjustments ──
    row = section_header(ws, row, 2, 6, '  SECTION D — Member Adjustments (Debit / Credit)')
    note_cell(ws, row, 2, 6, 'Use "debit" to add charge, "credit" to reduce amount for a member.')
    row += 1

    row_height(ws, row, 26)
    for i, h_ in enumerate(['Member Name', 'Type (debit/credit)', 'Amount (BDT)', 'Label / Reason'], 2):
        col_header(ws, row, i, h_)
    row += 1

    # Dropdown for type
    dv_adj = DataValidation(type='list', formula1='"debit,credit"', allow_blank=True)
    ws.add_data_validation(dv_adj)

    for i in range(5):
        row_height(ws, row, 22)
        bg = C['slate_50'] if i % 2 == 0 else C['white']
        cell(ws, row, 2, '', bg=bg, size=10, border=thin_border())
        c_ = cell(ws, row, 3, '', bg=bg, size=10, h='center', border=thin_border())
        dv_adj.add(c_)
        cell(ws, row, 4, 0, bg=bg, size=10, h='center', border=thin_border())
        cell(ws, row, 5, '', bg=bg, size=10, border=thin_border())
        row += 1

    return ws

# ── Sheet 4: Meals ────────────────────────────────────────────────────────────

def build_meals(wb):
    ws = wb.create_sheet('🍽 Meals')
    ws.sheet_view.showGridLines = False

    # Dynamic: 2 meal slots (Lunch, Dinner), up to 31 days
    # Layout: Col A=spacer, Col B=Member, Col C..=Day1L, Day1D, Day2L, Day2D ...
    # Each day occupies 2 columns (Lunch, Dinner)

    days = 31
    meal_slots = ['L', 'D']  # Lunch, Dinner
    members = SAMPLE_MEMBERS

    # Column widths
    set_col_width(ws, 'A', 3)
    set_col_width(ws, 'B', 22)
    for d in range(days):
        for s in range(len(meal_slots)):
            col_idx = 3 + d * len(meal_slots) + s
            set_col_width(ws, get_column_letter(col_idx), 4.5)

    # Shopping section starts after meal grid
    shop_col_start = 3 + days * len(meal_slots) + 2

    # Banner
    for c_ in range(1, 10):
        ws.cell(row=1, column=c_).fill = fill(C['slate_900'])
    row_height(ws, 1, 10)

    merge_cell(ws, 2, 2, 2, 12, '🍽  Meals — Monthly Checklist',
               bg=C['slate_900'], bold=True, size=15, color=C['amber'], h='left')
    row_height(ws, 2, 34)
    merge_cell(ws, 3, 2, 3, 12,
               'Mark ✓ (or V) for confirmed meals. L = Lunch, D = Dinner. Leave blank = not taken.',
               bg=C['amber'] + '22', bold=False, size=9, color=C['slate_500'], h='left', italic=True)
    row_height(ws, 3, 18)

    # ── Meal Grid Header ──
    row = 5
    row_height(ws, row, 20)
    cell(ws, row, 2, 'MEMBER', bg=C['slate_800'], bold=True, size=9,
         color=C['brand_light'], h='center', border=thin_border())

    # Day headers with merged cells per day
    for d in range(1, days + 1):
        col_idx = 3 + (d - 1) * len(meal_slots)
        ws.merge_cells(start_row=row, start_column=col_idx,
                        end_row=row, end_column=col_idx + len(meal_slots) - 1)
        c_ = ws.cell(row=row, column=col_idx, value=str(d))
        c_ .fill = fill(C['slate_700'] if d % 2 == 0 else C['slate_800'])
        c_.font = font(bold=True, size=8, color=C['slate_300'])
        c_.alignment = align(h='center', v='center')

    # Meal slot sub-headers
    row += 1
    row_height(ws, row, 18)
    cell(ws, row, 2, '', bg=C['slate_800'], border=thin_border())
    for d in range(days):
        for s_idx, slot in enumerate(meal_slots):
            col_idx = 3 + d * len(meal_slots) + s_idx
            c_ = ws.cell(row=row, column=col_idx, value=slot)
            c_.fill = fill(C['amber'] + '55' if slot == 'L' else C['sky'] + '44')
            c_.font = font(bold=True, size=7, color=C['slate_900'])
            c_.alignment = align(h='center', v='center')
            c_.border = thin_border()

    # Member rows
    for m_idx, member in enumerate(members):
        row += 1
        row_height(ws, row, 20)
        bg_row = C['slate_50'] if m_idx % 2 == 0 else C['white']
        cell(ws, row, 2, member, bg=bg_row, size=9, bold=True,
             color=C['slate_700'], border=thin_border())
        for d in range(days):
            for s_idx in range(len(meal_slots)):
                col_idx = 3 + d * len(meal_slots) + s_idx
                c_ = ws.cell(row=row, column=col_idx, value='')
                c_.fill = fill(bg_row)
                c_.font = font(size=9, color=C['brand'], bold=True)
                c_.alignment = align(h='center', v='center')
                c_.border = thin_border()

    # Blank extra member rows
    for i in range(4):
        row += 1
        row_height(ws, row, 20)
        bg_row = C['slate_50'] if (len(members) + i) % 2 == 0 else C['white']
        cell(ws, row, 2, '', bg=bg_row, size=9, border=thin_border())
        for d in range(days):
            for s_idx in range(len(meal_slots)):
                col_idx = 3 + d * len(meal_slots) + s_idx
                c_ = ws.cell(row=row, column=col_idx, value='')
                c_.fill = fill(bg_row)
                c_.border = thin_border()

    # ── Shopping Section ──
    row += 2
    row = section_header(ws, row, 2, 10, '  MEAL SHOPPING — Grocery Purchases This Month')
    note_cell(ws, row, 2, 10, 'Each row is one shopping trip. Amount = total spent (BDT integers).')
    row += 1

    row_height(ws, row, 26)
    shop_headers = ['#', 'Member Name', 'Item / Description', 'Amount (BDT)', 'Purchase Date (DD/MM/YYYY)', 'Notes']
    for i, h_ in enumerate(shop_headers, 2):
        col_header(ws, row, i, h_)
    row += 1

    for i in range(15):
        row_height(ws, row, 22)
        bg = C['slate_50'] if i % 2 == 0 else C['white']
        cell(ws, row, 2, i + 1, bg=bg, size=10, h='center', border=thin_border())
        cell(ws, row, 3, SAMPLE_MEMBERS[i % len(SAMPLE_MEMBERS)] if i < 8 else '', bg=bg, size=10, border=thin_border())
        cell(ws, row, 4, 'Groceries from market' if i < 3 else '', bg=bg, size=10, border=thin_border())
        c_ = ws.cell(row=row, column=5, value=1200 if i < 3 else 0)
        c_.fill = fill(bg)
        c_.font = font(size=10, color=C['amber'], bold=True)
        c_.alignment = align(h='center')
        c_.border = thin_border()
        c_.number_format = '#,##0'
        cell(ws, row, 6, '', bg=bg, size=10, h='center', border=thin_border())
        cell(ws, row, 7, '', bg=bg, size=10, border=thin_border())
        row += 1

    return ws

# ── Sheet 5: Expenses ─────────────────────────────────────────────────────────

def build_expenses(wb):
    ws = wb.create_sheet('💳 Expenses')
    ws.sheet_view.showGridLines = False

    widths = [3, 6, 22, 30, 16, 22, 22, 22, 3]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, get_column_letter(i), w)

    # Banner
    for c_ in range(1, 10):
        ws.cell(row=1, column=c_).fill = fill(C['slate_900'])
    row_height(ws, 1, 10)

    merge_cell(ws, 2, 2, 2, 8, '💳  Personal & Shared Expenses',
               bg=C['slate_900'], bold=True, size=15, color=C['rose'], h='left')
    row_height(ws, 2, 34)
    merge_cell(ws, 3, 2, 3, 8,
               'Log every expense item. One row = one expense. Date format: DD/MM/YYYY.',
               bg=C['rose'] + '22', bold=False, size=9, color=C['slate_500'], h='left', italic=True)
    row_height(ws, 3, 18)

    # Category legend
    row = 5
    row_height(ws, row, 22)
    merge_cell(ws, row, 2, row, 8, 'CATEGORY REFERENCE',
               bg=C['slate_800'], bold=True, size=10, color=C['rose'], h='left')
    row += 1

    cat_colors = {
        'Food': C['food'], 'Groceries': C['groceries'], 'Utilities': C['utilities'],
        'Transport': C['transport'], 'Household': C['household'],
        'Entertainment': C['entertainment'], 'Medical': C['medical'], 'Other': C['other'],
    }
    row_height(ws, row, 20)
    col = 2
    for cat, color in cat_colors.items():
        c_ = ws.cell(row=row, column=col, value=cat)
        c_.fill = fill(color + '44')
        c_.font = font(bold=True, size=9, color=C['slate_800'])
        c_.alignment = align(h='center', v='center')
        c_.border = thin_border()
        col += 1

    row += 2

    # Headers
    row_height(ws, row, 26)
    headers = ['#', 'Member Name', 'Item / Description', 'Amount (BDT)', 'Category', 'Date (DD/MM/YYYY)', 'Notes']
    for i, h_ in enumerate(headers, 2):
        col_header(ws, row, i, h_)
    row += 1

    # Category dropdown
    cat_list = ','.join(EXPENSE_CATEGORIES)
    dv_cat = DataValidation(type='list',
                             formula1=f'"{cat_list}"',
                             allow_blank=True,
                             showErrorMessage=True,
                             error='Choose from the list',
                             errorTitle='Invalid Category')
    ws.add_data_validation(dv_cat)

    # Sample data + blank rows
    sample_expenses = [
        ('Ali Hassan', 'Grocery shopping', 850, 'Groceries', '01/06/2025'),
        ('Rahim Uddin', 'Internet bill', 800, 'Utilities', '05/06/2025'),
        ('Karim Molla', 'Uber ride', 150, 'Transport', '07/06/2025'),
        ('Sumon Ahmed', 'Medicine', 350, 'Medical', '10/06/2025'),
        ('Ali Hassan', 'Cleaning supplies', 420, 'Household', '12/06/2025'),
    ]

    for i in range(40):
        row_height(ws, row, 22)
        bg = C['slate_50'] if i % 2 == 0 else C['white']

        cell(ws, row, 2, i + 1, bg=bg, size=10, h='center', border=thin_border())

        if i < len(sample_expenses):
            member, item, amount, cat, date = sample_expenses[i]
            cell(ws, row, 3, member, bg=bg, size=10, border=thin_border())
            cell(ws, row, 4, item, bg=bg, size=10, border=thin_border())

            cat_color = cat_colors.get(cat, C['other'])
            price_cell = ws.cell(row=row, column=5, value=amount)
            price_cell.fill = fill(bg)
            price_cell.font = font(size=10, color=C['rose'], bold=True)
            price_cell.alignment = align(h='center')
            price_cell.border = thin_border()
            price_cell.number_format = '#,##0'

            c_cat = ws.cell(row=row, column=6, value=cat)
            c_cat.fill = fill(cat_color + '33')
            c_cat.font = font(size=9, bold=True, color=C['slate_700'])
            c_cat.alignment = align(h='center')
            c_cat.border = thin_border()
            dv_cat.add(c_cat)

            cell(ws, row, 7, date, bg=bg, size=10, h='center', border=thin_border())
            cell(ws, row, 8, '', bg=bg, size=10, border=thin_border())
        else:
            for col in range(3, 9):
                c_ = ws.cell(row=row, column=col, value='')
                c_.fill = fill(bg)
                c_.border = thin_border()
                if col == 6:
                    dv_cat.add(c_)

        row += 1

    # Total row
    row_height(ws, row, 26)
    merge_cell(ws, row, 2, row, 4, 'TOTAL',
               bg=C['slate_800'], bold=True, size=11, color=C['white'], h='right')
    total_start = row - 40
    total_formula = f'=SUM(E{total_start}:E{row-1})'
    c_ = ws.cell(row=row, column=5, value=total_formula)
    c_.fill = fill(C['slate_800'])
    c_.font = font(bold=True, size=12, color=C['rose'])
    c_.alignment = align(h='center')
    c_.border = thin_border()
    c_.number_format = '#,##0 "৳"'

    return ws

# ── Build & Save ─────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb)
    build_members(wb)
    build_bills(wb)
    build_meals(wb)
    build_expenses(wb)

    # Tab colors
    tab_colors = {
        '📋 Instructions': '14B8A6',
        '👥 Members':      '22C55E',
        '🧾 Bills':        '38BDF8',
        '🍽 Meals':        'F59E0B',
        '💳 Expenses':     'FB7185',
    }
    for ws in wb.worksheets:
        color = tab_colors.get(ws.title, '94A3B8')
        ws.sheet_properties.tabColor = color
        ws.freeze_panes = ws['B5']

    out_path = os.path.join(
        os.path.dirname(__file__), '..', 'public', 'assets',
        'LocalHost_Monthly_Template.xlsx'
    )
    out_path = os.path.normpath(out_path)
    wb.save(out_path)
    print(f'✅  Saved: {out_path}')

if __name__ == '__main__':
    main()
